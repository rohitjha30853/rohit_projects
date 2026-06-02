# mmm_agg_qa.py
# pip install flask pandas openpyxl numpy
# python mmm_agg_qa.py
#
# What's new:
#   * After upload you land on a REVIEW page that pairs up the headers that
#     exist in only one file. If an R-file header was renamed in the JS file
#     (e.g. performance_max__unknown__none__none__spend  -> performance__max__
#     google__ads__none__non_brand__spend) you map them so they collapse into a
#     single comparison instead of two mystery lists.
#   * Auto-suggestions: a pair is pre-selected when the two headers' DAILY VALUE
#     SERIES line up across shared dates, or when the built-in HEADER_ALIASES
#     dict already knows the rename.
#   * Mapped headers are then compared by value like any shared column, and the
#     "Header/Features Comparison" section is rewritten to show:
#         - Renamed / Mapped Features : original -> renamed, how it matched, do values agree
#         - Unmatched Headers         : the genuine extras left in only one file

from flask import Flask, request, render_template_string, send_file, flash, redirect, url_for
import pandas as pd
import numpy as np
import re, io, uuid
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = "change-me"

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

TOLERANCE = 0.01
FEATID_RE = re.compile(r"^(featid\d+)__(.+)$")

HEADER_ALIASES = {
    "paid__search__google__ads__none__non_brand__impressions": "paid_search__google_ads__none__nonbrand__impressions",
    "paid__search__google__ads__none__non_brand__spend":       "paid_search__google_ads__none__nonbrand__spend",
    "paid__search__microsoft__ads__none__non_brand__impressions": "paid_search__microsoft_ads__none__nonbrand__impressions",
    "paid__search__microsoft__ads__none__non_brand__spend":       "paid_search__microsoft_ads__none__nonbrand__spend",
    "performance__max__google__ads__none__non_brand__impressions": "performance_max__unknown__none__none__impressions",
    "performance__max__google__ads__none__non_brand__spend":       "performance_max__unknown__none__none__spend",
    "podcast__all__none__none__impressions": "podcast__total__none__none__impressions",
    "podcast__all__none__none__spend":       "podcast__total__none__none__spend",
}

# In-memory store so the review page can regenerate without re-uploading.
SESSION_CACHE = OrderedDict()
_CACHE_LIMIT = 20

def _cache_put(payload: dict) -> str:
    token = uuid.uuid4().hex
    SESSION_CACHE[token] = payload
    while len(SESSION_CACHE) > _CACHE_LIMIT:
        SESSION_CACHE.popitem(last=False)
    return token

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def load_file(file_stream, filename: str, sheet_name: str = None) -> pd.DataFrame:
    data = io.BytesIO(file_stream.read())
    name = filename.lower()
    if name.endswith(".csv"):
        return pd.read_csv(data)
    elif name.endswith((".xlsx", ".xls")):
        xls = pd.ExcelFile(data, engine="openpyxl")
        target = sheet_name.strip() if sheet_name and sheet_name.strip() else xls.sheet_names[0]
        if target not in xls.sheet_names:
            raise ValueError(f"Sheet '{target}' not found. Available: {xls.sheet_names}")
        return xls.parse(target)
    raise ValueError(f"Unsupported file: '{filename}'. Use .csv or .xlsx.")

def load_file_for_qa(file_stream, filename: str, sheet_name: str = None) -> pd.DataFrame:
    """Like load_file but also normalises/finds the date column."""
    df = load_file(file_stream, filename, sheet_name)
    df.columns = [str(c) for c in df.columns]
    date_col = next((c for c in df.columns if c.strip().lower() == "date"), df.columns[0])
    df = df.rename(columns={date_col: "date"})
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.dropna(subset=["date"]).drop_duplicates(subset=["date"])
    return df

# ---------------------------------------------------------------------------
# featid Mapping logic
# ---------------------------------------------------------------------------

def apply_mapping(df: pd.DataFrame, mapping: dict) -> pd.DataFrame:
    orig_cols = list(df.columns)

    def rename_col(col):
        m = FEATID_RE.match(str(col))
        if not m:
            return col
        featid, suffix = m.group(1), m.group(2)
        return f"{mapping.get(featid, featid)}__{suffix}"

    df = df.copy()
    df.columns = [rename_col(c) for c in orig_cols]

    featid_slots = [i for i, c in enumerate(orig_cols) if FEATID_RE.match(str(c))]
    sorted_names = sorted(df.columns[i] for i in featid_slots)
    new_order = list(df.columns)
    for slot, name in zip(featid_slots, sorted_names):
        new_order[slot] = name
    return df[new_order]

# ---------------------------------------------------------------------------
# Header canonicalisation
# ---------------------------------------------------------------------------

def normalize(h: str) -> str:
    if h is None:
        return ""
    s = str(h).strip().lower()
    return re.sub(r"_+", "_", s)

def canonical(h: str) -> str:
    # Pure normalisation. Known renames are surfaced as *suggestions* in the
    # review step (via ALIAS_NORM) rather than silently merged here.
    return normalize(h)

# normalized-JS-header -> normalized-R-header, derived from the alias dict
ALIAS_NORM = {normalize(k): normalize(v) for k, v in HEADER_ALIASES.items()}

# ---------------------------------------------------------------------------
# Value-series matching (used for auto-suggestions and pair labelling)
# ---------------------------------------------------------------------------

def _series_vector(idx_df: pd.DataFrame, header: str, dates) -> np.ndarray:
    s = pd.to_numeric(idx_df.loc[dates, header], errors="coerce")
    return s.to_numpy(dtype=float)

def _vectors_match(va: np.ndarray, vb: np.ndarray, tol=TOLERANCE) -> bool:
    if va.shape != vb.shape or va.size == 0:
        return False
    nan_a, nan_b = np.isnan(va), np.isnan(vb)
    if (nan_a ^ nan_b).any():          # NaN vs number -> not a match
        return False
    both_nan = nan_a & nan_b
    denom = np.maximum(np.abs(va), np.abs(vb))
    with np.errstate(divide="ignore", invalid="ignore"):
        rel = np.where(denom == 0, 0.0, np.abs(va - vb) / denom)
    rel = np.where(both_nan, 0.0, rel)
    return bool(np.all(rel <= tol))

# ---------------------------------------------------------------------------
# Core comparison
# ---------------------------------------------------------------------------

def compute_discrepancies(a: pd.DataFrame, b: pd.DataFrame, header_map: dict = None,
                          start: pd.Timestamp = None, end: pd.Timestamp = None) -> dict:
    """header_map: {R header (original) -> JS header (renamed)}.
    start/end: optional inclusive date range. The whole QA is restricted to it,
    and the Date Discrepancies report lists, per file, the daily dates within the
    range that are absent from that file."""
    A, B = "JS File", "R File"

    # normalise dates to day-granularity and restrict to the selected range
    a = a.copy(); b = b.copy()
    a["date"] = pd.to_datetime(a["date"], errors="coerce").dt.normalize()
    b["date"] = pd.to_datetime(b["date"], errors="coerce").dt.normalize()
    a = a.dropna(subset=["date"]).drop_duplicates(subset=["date"])
    b = b.dropna(subset=["date"]).drop_duplicates(subset=["date"])
    if start is not None:
        a = a[a["date"] >= start]; b = b[b["date"] >= start]
    if end is not None:
        a = a[a["date"] <= end];   b = b[b["date"] <= end]

    a_map, b_map = {}, {}
    for c in a.columns:
        if c == "date": continue
        a_map.setdefault(canonical(c), c)
    for c in b.columns:
        if c == "date": continue
        b_map.setdefault(canonical(c), c)

    dates_a, dates_b = set(a["date"]), set(b["date"])
    shared_dates = sorted(dates_a & dates_b)
    a_i = a.set_index("date")
    b_i = b.set_index("date")

    # Expected daily calendar within the range -> dates missing from each file.
    # When start/end are not supplied, the calendar spans the data's own extent.
    union = dates_a | dates_b
    cal_start = start if start is not None else (min(union) if union else None)
    cal_end   = end   if end   is not None else (max(union) if union else None)
    if cal_start is not None and cal_end is not None and cal_start <= cal_end:
        expected = set(pd.date_range(cal_start, cal_end, freq="D"))
    else:
        expected = set()
    missing_a = sorted(expected - dates_a)   # days in range absent from JS File
    missing_b = sorted(expected - dates_b)   # days in range absent from R File

    # --- apply header mappings -------------------------------------------------
    # header_map is {R header (original) -> JS header (renamed)}. Several R
    # headers may point at the SAME JS header, in which case the R columns are
    # SUMMED before the comparison. e.g. on the R side
    #     paid_search__apple_search_ads__b_non_core__none__spend
    #   + paid_search__apple_search_ads__nb__none__spend
    # collapse into the single JS column
    #     paid_search__apple_search_ads__none__none__spend
    pair_meta = OrderedDict()   # canon -> (r_header(s), js_header, basis)
    if header_map:
        # group the mapped R headers by the JS header they target
        js_to_rs = OrderedDict()
        for r_header, js_header in header_map.items():
            if not r_header or not js_header:
                continue
            js_to_rs.setdefault(js_header, []).append(r_header)

        for js_header, r_headers in js_to_rs.items():
            ca = canonical(js_header)
            # the JS side must be a JS-only column we haven't already consumed
            if ca not in a_map or ca in b_map:
                continue
            # keep only the R headers that are genuine, distinct R-only columns
            seen, valid_rs = set(), []
            for rh in r_headers:
                cb = canonical(rh)
                if cb in b_map and cb not in a_map and cb != ca and cb not in seen:
                    valid_rs.append(rh)
                    seen.add(cb)
            if not valid_rs:
                continue

            if len(valid_rs) == 1:
                # ---- 1:1 rename: unify the JS canon onto the R canon ----
                r_header = valid_rs[0]
                cb = canonical(r_header)
                if ALIAS_NORM.get(ca) == cb:
                    basis = "Alias"
                else:
                    try:
                        va = _series_vector(a_i, a_map[ca], shared_dates)
                        vb = _series_vector(b_i, b_map[cb], shared_dates)
                        basis = "Values" if _vectors_match(va, vb) else "Manual"
                    except Exception:
                        basis = "Manual"
                js_col = a_map.pop(ca)
                a_map[cb] = js_col            # now both maps share canon cb
                pair_meta[cb] = (b_map[cb], js_col, basis)
            else:
                # ---- many R -> one JS: sum the R columns, compare to the JS one ----
                r_cols = [b_map[canonical(rh)] for rh in valid_rs]
                synth_name = f"__sum__{a_map[ca]}"
                num = b_i[r_cols].apply(lambda s: pd.to_numeric(s, errors="coerce"))
                b_i[synth_name] = num.sum(axis=1, min_count=1)
                # register the summed series under the JS canon so it lines up,
                # and drop the individual R-only canons so they aren't "unmatched"
                for rh in valid_rs:
                    b_map.pop(canonical(rh), None)
                b_map[ca] = synth_name        # ca now shared (JS already holds it)
                pair_meta[ca] = (" + ".join(r_cols), a_map[ca], "Sum")

    only_a = sorted(set(a_map) - set(b_map))
    only_b = sorted(set(b_map) - set(a_map))
    shared = sorted(set(a_map) & set(b_map))

    # --- value comparison over shared columns (now includes mapped pairs) ----
    mismatches = []
    mismatch_canons = set()
    for canon in shared:
        ah, bh = a_map[canon], b_map[canon]
        sa = pd.to_numeric(a_i.loc[shared_dates, ah], errors="coerce")
        sb = pd.to_numeric(b_i.loc[shared_dates, bh], errors="coerce")
        denom = np.maximum(np.abs(sa), np.abs(sb))
        with np.errstate(divide="ignore", invalid="ignore"):
            rel = np.where(denom == 0, 0.0, np.abs(sa - sb) / denom)
        mask = rel > TOLERANCE
        if mask.any():
            mismatch_canons.add(canon)
        for d, va, vb, rr in zip(np.array(shared_dates)[mask], sa[mask], sb[mask], rel[mask]):
            mismatches.append((pd.Timestamp(d), canon, ah, bh, va, vb, rr))

    # --- mapped-pair report rows -------------------------------------------
    mapped_pairs = []
    for cb, (rH, jH, basis) in pair_meta.items():
        agree = "N" if cb in mismatch_canons else "Y"
        mapped_pairs.append((rH, jH, basis, agree))

    return {
        "A": A, "B": B,
        "a_f": a, "b_f": b,
        "a_map": a_map, "b_map": b_map,
        "only_a": only_a, "only_b": only_b, "shared": shared,
        "missing_a": missing_a, "missing_b": missing_b,
        "shared_dates": shared_dates,
        "cal_start": cal_start, "cal_end": cal_end,
        "mismatches": mismatches,
        "mapped_pairs": mapped_pairs,
    }

# ---------------------------------------------------------------------------
# Suggestions for the review page
# ---------------------------------------------------------------------------

def build_suggestions(base: dict):
    """For each R-only header (original), suggest a JS-only header (renamed) by
    matching daily values, or via the alias dict. Returns (anchor_items, target_items)."""
    a_i = base["a_f"].set_index("date")
    b_i = base["b_f"].set_index("date")
    dates = base["shared_dates"]

    only_a, only_b = base["only_a"], base["only_b"]
    a_map, b_map   = base["a_map"], base["b_map"]

    js_vec = {}
    for ca in only_a:
        try:
            js_vec[ca] = _series_vector(a_i, a_map[ca], dates)
        except Exception:
            js_vec[ca] = np.array([])

    target_items = [{"name": a_map[ca]} for ca in only_a]

    used = set()
    anchor_items = []
    for i, cb in enumerate(only_b):
        rb = b_map[cb]
        try:
            vb = _series_vector(b_i, rb, dates)
        except Exception:
            vb = np.array([])
        suggestion, basis = "", ""

        # 1) alias-known rename
        for ca in only_a:
            if ca in used:
                continue
            if ALIAS_NORM.get(ca) == cb:
                suggestion, basis = a_map[ca], "alias"
                used.add(ca)
                break
        # 2) value-series match
        if not suggestion:
            for ca in only_a:
                if ca in used:
                    continue
                if _vectors_match(js_vec.get(ca, np.array([])), vb):
                    suggestion, basis = a_map[ca], "values"
                    used.add(ca)
                    break

        anchor_items.append({"idx": i, "name": rb, "suggestion": suggestion, "basis": basis})

    return anchor_items, target_items

# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def build_workbook(res: dict, client_name: str = "") -> io.BytesIO:
    A, B = res["A"], res["B"]
    a, b = res["a_f"], res["b_f"]
    only_a, only_b = res["only_a"], res["only_b"]
    a_map, b_map = res["a_map"], res["b_map"]
    mapped_pairs = res["mapped_pairs"]
    missing_a, missing_b = res["missing_a"], res["missing_b"]
    shared_dates, shared = res["shared_dates"], res["shared"]
    mismatches = res["mismatches"]
    cal_start, cal_end = res["cal_start"], res["cal_end"]
    range_label = (f"{cal_start:%m/%d/%Y} → {cal_end:%m/%d/%Y}"
                   if cal_start is not None and cal_end is not None else "(all dates)")

    wb  = Workbook()
    bold      = Font(bold=True)
    big       = Font(bold=True, size=14)
    head_fill = PatternFill("solid", fgColor="D9E1F2")
    ok_fill   = PatternFill("solid", fgColor="E2EFDA")
    bad_fill  = PatternFill("solid", fgColor="FCE4D6")

    def _write_df(ws, df):
        for ci, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = bold; cell.fill = head_fill
        for ri, vals in enumerate(df.values.tolist(), 2):
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if isinstance(val, pd.Timestamp):
                    cell.number_format = "m/d/yyyy"
        for ci, col in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(str(col)), 18)

    def _safe(name):
        return re.sub(r'[\\/*?:\[\]]', '_', name)[:31]

    ta = _safe(A); tb = _safe(B)
    if tb == ta:
        tb = _safe(f"{B} (2)")
    ws_a = wb.create_sheet(title=ta, index=0)
    ws_b = wb.create_sheet(title=tb, index=1)
    _write_df(ws_a, a)
    _write_df(ws_b, b)

    # ---- helper: description banner merged across the top of a sheet ----
    desc_font = Font(italic=True, size=13, color="FF555555")
    def _write_description(ws, text, ncols, total_width):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        cell = ws.cell(row=1, column=1, value=text)
        cell.font = desc_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cpl = max(int(total_width), 10)            # approx chars that fit on one line
        lines = max(1, (len(text) + cpl - 1) // cpl)
        ws.row_dimensions[1].height = lines * 17 + 8

    # =====================================================================
    # Sheet 3: Summary & Header/Features Comparison
    # =====================================================================
    ws = wb.create_sheet(title="Summary & Header Comparison", index=2)
    _write_description(
        ws,
        "Overview of the QA comparison between the two files, plus a header/features "
        "comparison: which features were renamed/mapped between the files and which "
        "headers appear in only one file.",
        ncols=2, total_width=100
    )

    r = 3
    # --- Summary ---
    ws.cell(row=r, column=1, value="Summary").font = big
    r += 1
    summary_rows = [
        ("Client name", client_name),
        ("QA date range", range_label),
        ("Renamed headers mapped (R\u2192JS)", len(mapped_pairs)),
        (f"Headers only in {A}", len(only_a)),
        (f"Headers only in {B}", len(only_b)),
        (f"Dates missing in {A} (within range)", len(missing_a)),
        (f"Dates missing in {B} (within range)", len(missing_b)),
        ("Shared dates compared", len(shared_dates)),
        ("Shared columns compared (incl. mapped)", len(shared)),
    ]
    left = Alignment(horizontal="left")
    for k, v in summary_rows:
        c1 = ws.cell(row=r, column=1, value=k); c1.font = bold; c1.alignment = left
        c2 = ws.cell(row=r, column=2, value=v); c2.alignment = left
        r += 1

    # --- Header/Features Comparison ---
    r += 1
    ws.cell(row=r, column=1, value="Header/Features Comparison").font = big
    r += 1

    # (a) Renamed / mapped features
    ws.cell(row=r, column=1, value="Renamed / Mapped Features").font = bold
    r += 1
    map_hdrs = [f"{B} header (original)", f"{A} header (renamed)"]
    for ci, h in enumerate(map_hdrs, start=1):
        c = ws.cell(row=r, column=ci, value=h); c.font = bold; c.fill = head_fill
    r += 1
    if mapped_pairs:
        for rH, jH, basis, agree in mapped_pairs:
            fill = ok_fill if agree == "Y" else bad_fill
            for ci, val in enumerate((rH, jH), start=1):
                cell = ws.cell(row=r, column=ci, value=val); cell.fill = fill
            r += 1
    else:
        ws.cell(row=r, column=1, value="(none mapped)"); r += 1

    r += 1
    # (b) Genuinely unmatched headers
    ws.cell(row=r, column=1, value="Unmatched Headers").font = bold
    r += 1
    ws.cell(row=r, column=1, value=f"Header in {A} (only)").font = bold
    ws.cell(row=r, column=2, value=f"Header in {B} (only)").font = bold
    ws.cell(row=r, column=1).fill = head_fill; ws.cell(row=r, column=2).fill = head_fill
    r += 1
    if not only_a and not only_b:
        ws.cell(row=r, column=1, value="(none)"); r += 1
    else:
        for i in range(max(len(only_a), len(only_b), 1)):
            if i < len(only_a):
                c = ws.cell(row=r, column=1, value=a_map[only_a[i]]); c.fill = bad_fill
            if i < len(only_b):
                c = ws.cell(row=r, column=2, value=b_map[only_b[i]]); c.fill = bad_fill
            r += 1

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 50

    # =====================================================================
    # Sheet 4: Value Mismatches (grouped by date as a collapsible hierarchy)
    # =====================================================================
    ws = wb.create_sheet(title="Value Mismatches", index=3)
    _write_description(
        ws,
        f"Value mismatches where the relative difference between the two files exceeds "
        f"{TOLERANCE:.0%}, grouped by year, month, and date (range: {range_label}). "
        f"Use the outline controls on the left to expand a year or month down to the "
        f"individual dates.",
        ncols=6, total_width=165
    )

    r = 3
    ws.cell(row=r, column=1, value="Value Mismatches (>1% relative difference)").font = big
    r += 1

    hdrs = [
        "Year / Month / Date",
        f"Header in {A}",
        f"Header in {B}",
        f"Value in {A}",
        f"Value in {B}",
        "% Difference"
    ]

    for i, h in enumerate(hdrs, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = bold
        c.fill = head_fill

    r += 1

    ws.sheet_properties.outlinePr.summaryBelow = False

    year_fill = PatternFill("solid", fgColor="EDEFF7")
    month_fill = PatternFill("solid", fgColor="F4F6FB")

    def _indent(row_idx, level):
        ws.cell(row=row_idx, column=1).alignment = Alignment(
            horizontal="left",
            indent=level
        )

    def _write_group_summary(label, rows, fill, outline_level):
        """
        Writes ONE summary row PER COLUMN instead of concatenating all
        columns into a single cell.
        """
        nonlocal r

        by_column = OrderedDict()

        for row in rows:
            canon = row[1]
            by_column.setdefault(canon, []).append(row)

        first_row = True

        for canon, col_rows in by_column.items():

            sa = float(np.nansum([x[4] for x in col_rows]))
            sb = float(np.nansum([x[5] for x in col_rows]))

            denom = max(abs(sa), abs(sb))
            pct = (abs(sa - sb) / denom) if denom else 0.0

            ah = ", ".join(
                sorted({str(x[2]) for x in col_rows})
            )

            bh = ", ".join(
                sorted({str(x[3]) for x in col_rows})
            )

            if first_row:
                ws.cell(row=r, column=1, value=label).font = bold
                first_row = False
            else:
                ws.cell(row=r, column=1, value="")

            ws.cell(row=r, column=2, value=ah)
            ws.cell(row=r, column=3, value=bh)

            ws.cell(row=r, column=4, value=sa)
            ws.cell(row=r, column=4).number_format = "#,##0.00"

            ws.cell(row=r, column=5, value=sb)
            ws.cell(row=r, column=5).number_format = "#,##0.00"

            ws.cell(row=r, column=6, value=pct)
            ws.cell(row=r, column=6).number_format = "0.00%"

            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = fill

            _indent(r, outline_level)
            ws.row_dimensions[r].outline_level = outline_level

            if outline_level == 1:
                ws.row_dimensions[r].hidden = True
                ws.row_dimensions[r].collapsed = True

            r += 1

    years = OrderedDict()

    for d, canon, ah, bh, va, vb, rel in sorted(
        mismatches,
        key=lambda x: (x[0], x[1])
    ):
        d = pd.Timestamp(d)

        years.setdefault(d.year, OrderedDict()) \
             .setdefault(d.month, []) \
             .append((d, canon, ah, bh, va, vb, rel))

    if not years:
        ws.cell(row=r, column=1, value="(none)")
        r += 1

    else:

        for year, months in years.items():

            year_rows = [
                row
                for rows in months.values()
                for row in rows
            ]

            # YEAR LEVEL
            _write_group_summary(
                str(year),
                year_rows,
                year_fill,
                0
            )

            for month, rows in months.items():

                label = pd.Timestamp(
                    year=year,
                    month=month,
                    day=1
                ).strftime("%b %Y")

                # MONTH LEVEL
                _write_group_summary(
                    label,
                    rows,
                    month_fill,
                    2
                )

                # DATE DETAILS
                for d, canon, ah, bh, va, vb, rel in rows:

                    dc = ws.cell(row=r, column=1, value=d)
                    dc.number_format = "m/d/yyyy"

                    ws.cell(row=r, column=2, value=ah)
                    ws.cell(row=r, column=3, value=bh)

                    ws.cell(
                        row=r,
                        column=4,
                        value=float(va) if pd.notna(va) else None
                    ).number_format = "#,##0.00"

                    ws.cell(
                        row=r,
                        column=5,
                        value=float(vb) if pd.notna(vb) else None
                    ).number_format = "#,##0.00"

                    ws.cell(
                        row=r,
                        column=6,
                        value=float(rel)
                    ).number_format = "0.00%"

                    _indent(r, 3)
                    ws.row_dimensions[r].outline_level = 3
                    ws.row_dimensions[r].hidden = True
                    ws.row_dimensions[r].collapsed = True

                    r += 1

    widths = {1: 40, 2: 34, 3: 34, 4: 22, 5: 22, 6: 13}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Collapse all groups by default
    for row_num, row_dim in ws.row_dimensions.items():

        if row_dim.outline_level >= 2:
            row_dim.hidden = True
            row_dim.collapsed = True

    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_view.showOutlineSymbols = True

    # =====================================================================
    # Sheet 5: Date Discrepancies (missing dates per file, within range)
    # =====================================================================
    ws = wb.create_sheet(title="Date Discrepancies", index=4)
    _write_description(
        ws,
        f"Dates within the selected QA range ({range_label}) that are missing from "
        f"each file. A date listed under a file means that file has no row for that day.",
        ncols=2, total_width=44
    )

    r = 3
    ws.cell(row=r, column=1, value=f"Date Discrepancies  (range: {range_label})").font = big
    r += 1
    ws.cell(row=r, column=1, value=f"Missing dates in {A}").font = bold
    ws.cell(row=r, column=2, value=f"Missing dates in {B}").font = bold
    ws.cell(row=r, column=1).fill = head_fill; ws.cell(row=r, column=2).fill = head_fill
    r += 1
    if not missing_a and not missing_b:
        ws.cell(row=r, column=1, value="(none \u2014 both files cover every day in range)"); r += 1
    else:
        for i in range(max(len(missing_a), len(missing_b))):
            if i < len(missing_a):
                c = ws.cell(row=r, column=1, value=missing_a[i]); c.number_format = "m/d/yyyy"; c.fill = bad_fill
            if i < len(missing_b):
                c = ws.cell(row=r, column=2, value=missing_b[i]); c.number_format = "m/d/yyyy"; c.fill = bad_fill
            r += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

# ---------------------------------------------------------------------------
# Flask UI
# ---------------------------------------------------------------------------

PAGE = r"""
<!doctype html><title>MMM Aggregated Daily QA</title>
<style>
body{font-family:system-ui;max-width:720px;margin:40px auto;padding:0 16px}
h1{margin-bottom:4px}
p.sub{color:#555;margin:0 0 16px}
input,button{font-size:15px;padding:8px;margin:4px 0}
label{display:block;margin:10px 0 2px;font-weight:600}
.hint{font-size:13px;color:#666;margin:0 0 2px;font-weight:normal}
.section{background:#f7f8fc;border:1px solid #dde;border-radius:8px;padding:16px 20px;margin:12px 0}
.section-title{font-size:16px;font-weight:700;margin:0 0 10px;color:#1a1a2e}
.step-badge{display:inline-block;background:#3b5bdb;color:#fff;border-radius:50%;
  width:22px;height:22px;text-align:center;line-height:22px;font-size:12px;
  font-weight:700;margin-right:6px;vertical-align:middle}
button[type=submit]{background:#3b5bdb;color:#fff;border:none;border-radius:6px;
  padding:10px 28px;cursor:pointer;margin-top:14px;font-size:15px}
button[type=submit]:hover{background:#2f4ac4}
.flash{background:#fee;border:1px solid #c00;border-radius:6px;padding:10px;margin:10px 0}
#paste-area{width:100%;box-sizing:border-box;font-family:monospace;font-size:13px;
  border:2px dashed #aab;border-radius:6px;padding:10px;min-height:72px;resize:vertical;
  background:#fff;color:#333;margin-top:4px}
#paste-area:focus{outline:none;border-color:#3b5bdb}
#parse-status{font-size:13px;margin:4px 0 0;min-height:16px;color:#2f9e44;font-weight:600}
#map-table{border-collapse:collapse;width:100%;margin-top:8px}
#map-table th{text-align:left;padding:6px 10px;background:#e8eaf6;font-size:13px;border:1px solid #ccd}
#map-table td{padding:3px 5px;border:1px solid #dde}
#map-table input{width:100%;box-sizing:border-box;border:1px solid #ccc;border-radius:4px;
  padding:4px 7px;font-size:13px;margin:0}
.btn-sm{border:none;border-radius:5px;color:#fff;cursor:pointer;font-size:13px;padding:5px 14px;margin-top:6px}
.btn-add{background:#2f9e44}.btn-add:hover{background:#258f38}
.btn-clr{background:#868e96;margin-left:6px}.btn-clr:hover{background:#6c757d}
.btn-del{background:#e03131;padding:3px 9px;font-size:12px;margin:0;border-radius:4px;border:none;color:#fff;cursor:pointer}
.btn-del:hover{background:#c92a2a}
</style>
<h1>MMM Aggregated Daily QA</h1>
<p class="sub">Compare two daily aggregated files. Rename <code>featidXXX</code> columns in the JavaScript Generated File before the comparison, then map any renamed headers on the next screen.</p>

{% with msgs = get_flashed_messages() %}{% for m in msgs %}<div class="flash">{{ m }}</div>{% endfor %}{% endwith %}

<form method="post" enctype="multipart/form-data" action="/run" id="main-form">

  <!-- Client name -->
  <div class="section">
    <div class="section-title">Client</div>
    <label>Client name <span class="hint" style="font-weight:400">(optional — shown in the report summary)</span></label>
    <input type="text" name="client_name" placeholder="e.g. Acme Corp" style="width:100%;box-sizing:border-box">
  </div>

  <!-- Step 1: Date range -->
  <div class="section">
    <div class="section-title"><span class="step-badge">1</span>Date range <span class="hint" style="font-weight:400">(optional)</span></div>
    <span class="hint">Restrict the QA to this range. Leave blank to use every date in the files. Missing dates are reported per file within this range.</span>
    <div style="display:flex;gap:18px;flex-wrap:wrap">
      <div><label>Start date</label><input type="date" name="start_date"></div>
      <div><label>End date</label><input type="date" name="end_date"></div>
    </div>
  </div>

  <!-- Step 2: JavaScript Generated File -->
  <div class="section">
    <div class="section-title"><span class="step-badge">2</span>JavaScript Generated File</div>
    <label>Upload file <span class="hint">(.csv or .xlsx)</span></label>
    <input type="file" name="file_a" accept=".csv,.xlsx,.xls" required>

    <div id="mapping-section" style="margin-top:14px">
      <label>Mapping
        <span class="hint">Paste two columns: <code>feature_id</code> and <code>feature_name</code>. Header row is skipped automatically.</span>
      </label>
      <textarea id="paste-area" placeholder="Paste mapping here (Ctrl+V / Cmd+V)&#10;&#10;feature_id&#9;feature_name&#10;featid001&#9;TV_Spend&#10;featid002&#9;Radio_Spend"></textarea>
      <div id="parse-status"></div>
      <table id="map-table">
        <thead><tr>
          <th style="width:40%">feature_id</th>
          <th style="width:50%">feature_name</th>
          <th style="width:10%"></th>
        </tr></thead>
        <tbody id="map-body"></tbody>
      </table>
      <button type="button" class="btn-sm btn-add" onclick="addRow()">+ Add row</button>
      <button type="button" class="btn-sm btn-clr" onclick="clearTable()">Clear</button>
      <input type="hidden" name="use_mapping" value="1" id="use-mapping-hidden">
    </div>
  </div>

  <!-- Step 3: R Script Generated File -->
  <div class="section">
    <div class="section-title"><span class="step-badge">3</span>R Script Generated File</div>
    <label>Upload file <span class="hint">(.csv or .xlsx)</span></label>
    <input type="file" name="file_b" accept=".csv,.xlsx,.xls" required>
  </div>

  <button type="submit">Run QA &amp; Review →</button>
</form>

<script>
function makeRow(id, name) {
  const tr = document.createElement('tr');
  tr.innerHTML =
    `<td><input type="text" name="feature_id" value="${esc(id)}" placeholder="e.g. featid001"></td>` +
    `<td><input type="text" name="feature_name" value="${esc(name)}" placeholder="e.g. TV_Spend"></td>` +
    `<td><button type="button" class="btn-del" onclick="removeRow(this)">&#x2715;</button></td>`;
  return tr;
}
function esc(v){return String(v).replace(/&/g,'&amp;').replace(/"/g,'&quot;').replace(/</g,'&lt;');}
function addRow(){document.getElementById('map-body').appendChild(makeRow('',''));}
function removeRow(btn){const t=document.getElementById('map-body');if(t.rows.length>1)btn.closest('tr').remove();}
function clearTable(){document.getElementById('map-body').innerHTML='';document.getElementById('parse-status').textContent='';addRow();}
function parsePaste(text){
  const lines=text.replace(/\r/g,'').split('\n').filter(l=>l.trim());
  if(!lines.length)return[];
  const first=lines[0];
  const delim=first.includes('\t')?'\t':first.includes(',')?',':null;
  const rows=lines.map(l=>{const cols=delim?l.split(delim):l.trim().split(/\s{2,}/);return cols.map(c=>c.trim().replace(/^["']+|["']+$/g,''));});
  const start=rows.length&&rows[0][0].toLowerCase().replace(/[^a-z_]/g,'')==='feature_id'?1:0;
  return rows.slice(start).filter(r=>r.length>=2&&r[0]&&r[1]);
}
function loadFromTextarea(){
  const ta=document.getElementById('paste-area');const status=document.getElementById('parse-status');const text=ta.value;
  if(!text.trim()){status.style.color='#c92a2a';status.textContent='Nothing to parse.';return false;}
  const rows=parsePaste(text);
  if(!rows.length){status.style.color='#c92a2a';status.textContent='Could not parse rows — check format (tab or comma separated, 2 columns).';return false;}
  const tbody=document.getElementById('map-body');tbody.innerHTML='';
  rows.forEach(r=>tbody.appendChild(makeRow(r[0],r[1])));
  ta.value='';status.style.color='#2f9e44';status.textContent='✓ Loaded '+rows.length+' row'+(rows.length!==1?'s':'');
  return true;
}
document.getElementById('paste-area').addEventListener('paste',function(e){
  e.preventDefault();const text=(e.clipboardData||window.clipboardData).getData('text');this.value=text;loadFromTextarea();
});
</script>
"""

REVIEW = r"""
<!doctype html><title>Review &amp; Map renamed headers</title>
<style>
body{font-family:system-ui;max-width:1040px;margin:36px auto;padding:0 16px;color:#222}
h1{margin-bottom:2px}
p.sub{color:#555;margin:0 0 18px}
.cards{display:flex;gap:14px;flex-wrap:wrap;margin:0 0 22px}
.card{flex:1;min-width:150px;background:#f7f8fc;border:1px solid #dde;border-radius:8px;padding:12px 16px}
.card .n{font-size:24px;font-weight:700}
.card .l{font-size:13px;color:#555}
table{border-collapse:collapse;width:100%;margin:8px 0 24px;font-size:13.5px}
th,td{border:1px solid #e2e2ec;padding:7px 9px;text-align:left;vertical-align:top}
th{background:#d9e1f2}
select{font-size:13px;padding:5px;max-width:430px}
.sugg{background:#fff7d6}
.tag{font-size:11px;font-weight:700;padding:1px 6px;border-radius:4px;margin-left:6px;color:#fff}
.tag.values{background:#2f9e44}.tag.alias{background:#7048e8}
.muted{color:#888}
.note{background:#eef4ff;border:1px solid #cdddff;border-radius:8px;padding:10px 14px;font-size:14px;margin:0 0 18px}
button{background:#3b5bdb;color:#fff;border:none;border-radius:6px;padding:11px 26px;cursor:pointer;font-size:15px}
button:hover{background:#2f4ac4}
a.back{display:inline-block;margin-left:14px;color:#3b5bdb;font-size:14px}
code{background:#f0f0f6;padding:1px 5px;border-radius:4px;font-size:12.5px}
h2{font-size:17px;margin:26px 0 6px}
</style>
<h1>Review &amp; map renamed headers</h1>
<p class="sub">Headers below exist in only one file. If an R-file header was <b>renamed</b> in the JS file, map it so the two collapse into a single comparison. &nbsp;<b>QA date range:</b> {{ range_label }}.</p>

<div class="cards">
  <div class="card"><div class="n">{{ n_shared }}</div><div class="l">Shared headers</div></div>
  <div class="card"><div class="n">{{ anchor_items|length }}</div><div class="l">Only in R File</div></div>
  <div class="card"><div class="n">{{ target_items|length }}</div><div class="l">Only in JS File</div></div>
  <div class="card"><div class="n">{{ n_suggested }}</div><div class="l">Auto-suggested</div></div>
</div>

<form method="post" action="/download">
  <input type="hidden" name="token" value="{{ token }}">

  <div class="note">
    Suggested matches are <span class="sugg">pre-selected</span> — by daily <b>values</b> lining up
    (<span class="tag values">values</span>) or a known <b>alias</b> (<span class="tag alias">alias</span>).
    Leave a row on <code>(no match — keep separate)</code> to keep it unmatched.
  </div>

  {% if anchor_items %}
  <h2>Map "Only in R File" → "Only in JS File"</h2>
  <table>
    <tr><th style="width:42%">R File header (original)</th><th>maps to → JS File header (renamed)</th></tr>
    {% for it in anchor_items %}
    <tr class="{{ 'sugg' if it.suggestion else '' }}">
      <td><code>{{ it.name }}</code>{% if it.basis %}<span class="tag {{ it.basis }}">{{ it.basis }}</span>{% endif %}</td>
      <td>
        <input type="hidden" name="r_{{ it.idx }}" value="{{ it.name }}">
        <select name="map_{{ it.idx }}">
          <option value="">(no match — keep separate)</option>
          {% for t in target_items %}
          <option value="{{ t.name }}" {{ 'selected' if t.name == it.suggestion else '' }}>{{ t.name }}</option>
          {% endfor %}
        </select>
      </td>
    </tr>
    {% endfor %}
  </table>
  {% else %}
  <p class="muted">No R-only headers to map.</p>
  {% endif %}

  {% if target_items %}
  <h2 class="muted">Reference: headers only in JS File</h2>
  <table>
    <tr><th>JS File header</th></tr>
    {% for t in target_items %}<tr><td><code>{{ t.name }}</code></td></tr>{% endfor %}
  </table>
  {% endif %}

  <button type="submit">Generate QA report ↓</button>
  <a class="back" href="/">↺ Start over</a>
</form>
"""

# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

def _parse_date(s):
    s = (s or "").strip()
    if not s:
        return None
    try:
        return pd.Timestamp(s).normalize()
    except Exception:
        return None

@app.route("/")
def index():
    return render_template_string(PAGE)

@app.route("/run", methods=["POST"])
def run():
    fa = request.files.get("file_a")
    fb = request.files.get("file_b")
    if not fa or not fa.filename:
        flash("Please upload JavaScript Generated File."); return redirect(url_for("index"))
    if not fb or not fb.filename:
        flash("Please upload R Script Generated File."); return redirect(url_for("index"))

    start = _parse_date(request.form.get("start_date"))
    end   = _parse_date(request.form.get("end_date"))
    if start is not None and end is not None and start > end:
        flash("Start date is after end date."); return redirect(url_for("index"))

    try:
        a = load_file_for_qa(fa.stream, fa.filename, request.form.get("sheet_a", ""))
        b = load_file_for_qa(fb.stream, fb.filename, request.form.get("sheet_b", ""))

        # Rename featid columns in File A
        if request.form.get("use_mapping", "0") == "1":
            feature_ids   = request.form.getlist("feature_id")
            feature_names = request.form.getlist("feature_name")
            mapping = {
                fid.strip(): fname.strip()
                for fid, fname in zip(feature_ids, feature_names)
                if fid.strip() and fname.strip()
            }
            if mapping:
                a = apply_mapping(a, mapping)

        base = compute_discrepancies(a, b, start=start, end=end)
        anchor_items, target_items = build_suggestions(base)
    except Exception as e:
        flash(f"Error: {e}"); return redirect(url_for("index"))

    token = _cache_put({"a": a, "b": b, "start": start, "end": end,
                        "client_name": request.form.get("client_name", "").strip()})
    n_suggested = sum(1 for it in anchor_items if it["suggestion"])
    range_label = (f"{base['cal_start']:%m/%d/%Y} → {base['cal_end']:%m/%d/%Y}"
                   if base["cal_start"] is not None and base["cal_end"] is not None else "(all dates)")
    return render_template_string(REVIEW, token=token,
                                  anchor_items=anchor_items, target_items=target_items,
                                  n_shared=len(base["shared"]), n_suggested=n_suggested,
                                  range_label=range_label)

@app.route("/download", methods=["POST"])
def download():
    token = request.form.get("token", "")
    payload = SESSION_CACHE.get(token)
    if not payload:
        flash("Session expired — please upload the files again.")
        return redirect(url_for("index"))

    # header_map: {R header (original) -> JS header (renamed)}
    header_map = {}
    for key, val in request.form.items():
        if key.startswith("map_") and val:
            idx = key[len("map_"):]
            r_header = request.form.get(f"r_{idx}", "")
            if r_header:
                header_map[r_header] = val

    try:
        res = compute_discrepancies(payload["a"], payload["b"], header_map=header_map,
                                    start=payload.get("start"), end=payload.get("end"))
        out = build_workbook(res, client_name=payload.get("client_name", ""))
    except Exception as e:
        flash(f"Error: {e}"); return redirect(url_for("index"))

    client_name = payload.get("client_name", "").strip()
    safe_client = re.sub(r"[^\w\-]", "_", client_name).lower() if client_name else ""
    filename = f"{safe_client}_mmm_aggregated_qa_summary.xlsx" if safe_client else "mmm_aggregated_qa_summary.xlsx"
    return send_file(out, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    app.run(debug=True, port=5000)