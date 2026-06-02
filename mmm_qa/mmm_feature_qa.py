# mmm_feature_qa.py
# pip install flask pandas openpyxl numpy
# python mmm_feature_qa.py
#
# What's new vs. the original:
#   * After upload you now land on a REVIEW page that lists the features that
#     only exist in one file ("Only in JS File" / "Only in R File").
#   * For genuine renames (e.g. performance_max_google_ads_none_non-brand  ->
#     performance_max__unknown__none__none_) you can map a JS feature to its
#     R-file counterpart with a dropdown. Pairs with identical spend are
#     pre-selected automatically as a suggestion.
#   * Mapped pairs collapse into a SINGLE row in the Comparison sheet, with two
#     new columns:
#         - feature_name_JS_to_R : holds both names on the same row ("old -> new")
#         - match_type           : Auto / Manual (manual maps are highlighted)

from flask import (Flask, request, render_template_string, send_file,
                   flash, redirect, url_for)
import pandas as pd
import numpy as np
import re, io, uuid
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = "change-me"

# Simple in-memory store so the review page can regenerate the report without
# re-uploading the files. Fine for a local single-user dev tool.
SESSION_CACHE = OrderedDict()
_CACHE_LIMIT = 20

def _cache_put(payload: dict) -> str:
    token = uuid.uuid4().hex
    SESSION_CACHE[token] = payload
    while len(SESSION_CACHE) > _CACHE_LIMIT:
        SESSION_CACHE.popitem(last=False)
    return token

# ---------------------------------------------------------------------------
# Feature key helpers
# ---------------------------------------------------------------------------

def _normalise_feature_key(name: str) -> str:
    """Collapse separators so  paid_search__google == paid_search_google."""
    if not isinstance(name, str):
        return name
    return re.sub(r"[_\-\s]+", "", name).lower()

# Known renames: {JS feature (renamed) -> R feature (original)} at the
# feature_og level (no _spend / _impressions suffix). These are surfaced as
# pre-selected *suggestions* on the review page (basis = "alias"); they are not
# silently merged. Names are compared via _normalise_feature_key, so separator
# and case differences don't matter. Pairs that only differ by separators
# (e.g. non_brand vs nonbrand) already auto-match and never reach the review
# page, so listing them here is harmless.
FEATURE_ALIASES = {
    "paid_search_google_ads_none_non_brand":     "paid_search_google_ads_none_nonbrand",
    "paid_search_microsoft_ads_none_non_brand":  "paid_search_microsoft_ads_none_nonbrand",
    "performance_max_google_ads_none_non_brand": "performance_max_unknown_none_none",
    "podcast_all_none_none":                      "podcast_total_none_none",
}

# normalized-JS-feature -> normalized-R-feature
ALIAS_NORM = {_normalise_feature_key(js): _normalise_feature_key(r)
              for js, r in FEATURE_ALIASES.items()}

def _strip_metric_suffix(name: str) -> str:
    if not isinstance(name, str):
        return name
    for suffix in ("_spend", "_impressions"):
        if name.endswith(suffix):
            return name[: -len(suffix)]
    return name

def _feature_og_from_formula(name: str) -> str:
    # =IFERROR(IF(RIGHT(B2,6)="_spend",LEFT(B2,LEN(B2)-7),IF(RIGHT(B2,12)="_impressions",LEFT(B2,LEN(B2)-13),B2)),B2)
    if not isinstance(name, str):
        return name
    try:
        if name[-6:] == "_spend":
            return name[:len(name) - 7]
        if name[-12:] == "_impressions":
            return name[:len(name) - 13]
        return name
    except Exception:
        return name

def _within(a, b, abs_tol, rel_tol) -> bool:
    if pd.isna(a) and pd.isna(b):
        return True
    if pd.isna(a) or pd.isna(b):
        return False
    diff = abs(a - b)
    if diff <= abs_tol:
        return True
    return diff / max(abs(a), abs(b), 1.0) <= rel_tol

# ---------------------------------------------------------------------------
# File loading
# ---------------------------------------------------------------------------

def load_file(file_stream, filename: str, sheet_name: str = None) -> pd.DataFrame:
    data = io.BytesIO(file_stream.read())
    name = filename.lower()
    if name.endswith(".csv"):
        return pd.read_csv(data)
    elif name.endswith((".xlsx", ".xls")):
        xls = pd.ExcelFile(data, engine="openpyxl")
        target = sheet_name.strip() if sheet_name and sheet_name.strip() else xls.sheet_names[0]
        if target not in xls.sheet_names:
            raise ValueError(f"Sheet '{target}' not found. Available: {xls.sheet_names}")
        return xls.parse(target)
    raise ValueError(f"Unsupported file type: '{filename}'. Use .csv or .xlsx.")

def prepare_dfs(qa_bytes, qa_filename, qa_sheet, pipe_bytes, pipe_filename, pipe_sheet):
    qa_df   = load_file(io.BytesIO(qa_bytes),   qa_filename,   qa_sheet)
    pipe_df = load_file(io.BytesIO(pipe_bytes), pipe_filename, pipe_sheet)
    if "mmm_feature_og" not in pipe_df.columns and "mmm_feature" in pipe_df.columns:
        idx = pipe_df.columns.get_loc("mmm_feature") + 1
        pipe_df.insert(idx, "mmm_feature_og", pipe_df["mmm_feature"].map(_feature_og_from_formula))
    return qa_df, pipe_df

# ---------------------------------------------------------------------------
# Normalisation
# ---------------------------------------------------------------------------

def normalise_qa(df: pd.DataFrame, feature_col="mmm_feature_og",
                 spend_col="spend", imp_col="impressions") -> pd.DataFrame:
    # The JS/QA file's feature column may be named differently. In some exports
    # the original feature lives in "feature_name" (which IS the mmm_feature_og
    # column), so accept the first of these that exists.
    if feature_col not in df.columns:
        for alt in ("feature_name", "mmm_feature_og", "mmm_feature"):
            if alt in df.columns:
                feature_col = alt
                break

    missing = [c for c in (feature_col, spend_col, imp_col) if c not in df.columns]
    if missing:
        raise KeyError(f"QA file missing columns: {missing}")
    work = df.copy()
    work[spend_col] = pd.to_numeric(work[spend_col], errors="coerce")
    work[imp_col]   = pd.to_numeric(work[imp_col],   errors="coerce")
    work["_key"] = work[feature_col].map(_normalise_feature_key)
    return (
        work.groupby(["_key", feature_col])
        .agg(spend_QA=(spend_col, "sum"), impressions_QA=(imp_col, "sum"),
             rows_QA=(feature_col, "size"))
        .reset_index()
        .rename(columns={feature_col: "mmm_feature_og_QA"})
    )

def normalise_pipeline(df: pd.DataFrame, feature_col="mmm_feature",
                       feature_og_col="mmm_feature_og",
                       value_col="value", type_col="original_type") -> pd.DataFrame:
    work = df.copy()

    # Derive mmm_feature_og if absent
    if feature_og_col not in work.columns:
        if feature_col not in work.columns:
            raise KeyError(f"Pipeline file needs '{feature_og_col}' or '{feature_col}'")
        work[feature_og_col] = work[feature_col].map(_strip_metric_suffix)

    # Derive type (spend/impressions) if absent
    if type_col not in work.columns:
        if feature_col not in work.columns:
            raise KeyError(f"Pipeline file needs '{type_col}' or '{feature_col}' to infer metric type")
        work[type_col] = np.where(
            work[feature_col].astype(str).str.endswith("_spend"), "spend",
            np.where(work[feature_col].astype(str).str.endswith("_impressions"), "impressions", "other")
        )

    if value_col not in work.columns:
        raise KeyError(f"Pipeline file needs '{value_col}'")

    work[value_col] = pd.to_numeric(work[value_col], errors="coerce")
    work["_key"] = work[feature_og_col].map(_normalise_feature_key)

    pivot = (
        work.pivot_table(index=["_key", feature_og_col], columns=type_col,
                         values=value_col, aggfunc="sum")
        .reset_index().rename_axis(None, axis=1)
    )
    counts = work.groupby("_key").size().rename("rows_pipeline").reset_index()
    out = pivot.merge(counts, on="_key", how="left")
    for col in ("spend", "impressions"):
        if col not in out.columns:
            out[col] = 0.0
    return out[["_key", feature_og_col, "spend", "impressions", "rows_pipeline"]]

# ---------------------------------------------------------------------------
# Core comparison
# ---------------------------------------------------------------------------

def compute_comparison(qa_df, pipe_df, manual_map=None,
                       spend_abs_tol=0.01, impressions_abs_tol=0.5):
    """Returns (features_df, summary).

    manual_map: dict {JS feature name -> R feature name}. Each mapped pair is
    forced to share a key so they merge into a single row and are flagged
    match_type == 'Manual'.

    Note: ``features`` keeps the internal helper columns ``status`` and
    ``mmm_feature_og`` for backend logic (sorting, review-page suggestions,
    row colouring). They are intentionally NOT written to the Comparison sheet.
    """
    qa_norm   = normalise_qa(qa_df)
    pipe_norm = normalise_pipeline(pipe_df).rename(columns={
        "spend":        "spend_Pipeline",
        "impressions":  "impressions_Pipeline",
        "mmm_feature_og": "mmm_feature_og_Pipeline",
    })

    # --- apply manual rename mappings ---------------------------------------
    # manual_map: {JS feature name -> R feature name OR list of R names}. When
    # several R features point at the same JS feature, their values are SUMMED
    # so the comparison sees a single R total. e.g. on the R side
    #     paid_search_apple_search_ads_b_non_core  (spend + impressions)
    #   + paid_search_apple_search_ads_nb          (spend + impressions)
    # collapse into the single JS feature paid_search_apple_search_ads_none_none.
    manual_keys = set()
    if manual_map:
        qa_key_by_name = dict(zip(qa_norm["mmm_feature_og_QA"], qa_norm["_key"]))
        used_r = set()
        for js_name, r_names in manual_map.items():
            if not js_name:
                continue
            jk = qa_key_by_name.get(js_name)
            if jk is None:
                continue
            if isinstance(r_names, str):
                r_names = [r_names]
            for r_name in r_names:
                if not r_name or r_name in used_r:
                    continue
                mask = pipe_norm["mmm_feature_og_Pipeline"] == r_name
                if mask.any():
                    pipe_norm.loc[mask, "_key"] = jk   # unify keys -> they merge
                    manual_keys.add(jk)
                    used_r.add(r_name)

    # Collapse any pipe-side rows that now share a key (many R -> one JS) into a
    # single summed row, so the merge sees one R total per feature instead of
    # one row per original R feature.
    if pipe_norm["_key"].duplicated().any():
        pipe_norm = (
            pipe_norm.groupby("_key", as_index=False).agg({
                "mmm_feature_og_Pipeline": lambda s: " + ".join(dict.fromkeys(s.astype(str))),
                "spend_Pipeline":          "sum",
                "impressions_Pipeline":    "sum",
                "rows_pipeline":           "sum",
            })
        )

    merged = qa_norm.merge(pipe_norm, on="_key", how="outer")
    merged["mmm_feature_og"] = merged["mmm_feature_og_QA"].fillna(merged["mmm_feature_og_Pipeline"])

    in_qa   = merged["rows_QA"].notna()   & (merged["rows_QA"].fillna(0) > 0)
    in_pipe = merged["rows_pipeline"].notna() & (merged["rows_pipeline"].fillna(0) > 0)
    merged["status"] = np.select(
        [in_qa & in_pipe, in_qa & ~in_pipe, ~in_qa & in_pipe],
        ["Both", "Only in JS File", "Only in R File"], default="Neither"
    )

    # Auto vs Manual match flag
    merged["match_type"] = np.where(
        (merged["status"] == "Both") & merged["_key"].isin(manual_keys), "Manual",
        np.where(merged["status"] == "Both", "Auto", "")
    )

    for col in ("spend_QA", "spend_Pipeline", "impressions_QA", "impressions_Pipeline"):
        merged[col] = merged[col].fillna(0.0)
    for col in ("rows_QA", "rows_pipeline"):
        merged[col] = merged[col].fillna(0).astype(int)

    merged["spend_diff"]      = merged["spend_Pipeline"]       - merged["spend_QA"]
    merged["impressions_diff"] = merged["impressions_Pipeline"] - merged["impressions_QA"]

    with np.errstate(divide="ignore", invalid="ignore"):
        merged["spend_pct_diff"] = np.where(
            merged["spend_QA"] != 0, merged["spend_diff"] / merged["spend_QA"],
            np.where(merged["spend_Pipeline"] == 0, 0.0, -1.0))
        merged["impressions_pct_diff"] = np.where(
            merged["impressions_QA"] != 0, merged["impressions_diff"] / merged["impressions_QA"],
            np.where(merged["impressions_Pipeline"] == 0, 0.0, -1.0))

    merged["spend_match"] = merged.apply(
        lambda r: "Y" if _within(r["spend_QA"], r["spend_Pipeline"], spend_abs_tol, 1e-6) else "N", axis=1)
    merged["imp_match"] = merged.apply(
        lambda r: "Y" if _within(r["impressions_QA"], r["impressions_Pipeline"], impressions_abs_tol, 1e-6) else "N", axis=1)

    # New consolidated column: both names on the same row, original -> renamed
    # (R File holds the original name, JS File holds the renamed one).
    def _combine(r):
        js = r["mmm_feature_og_QA"]
        rr = r["mmm_feature_og_Pipeline"]
        js = js if isinstance(js, str) else None
        rr = rr if isinstance(rr, str) else None
        if js and rr:
            return js if js == rr else f"{rr}  →  {js}"
        return rr or js or ""
    merged["feature_name_R_to_JS"] = merged.apply(_combine, axis=1)

    features = merged[[
        "status", "match_type", "spend_match", "imp_match",
        "mmm_feature_og", "mmm_feature_og_QA", "mmm_feature_og_Pipeline",
        "feature_name_R_to_JS",
        "rows_QA", "rows_pipeline",
        "spend_QA", "spend_Pipeline", "spend_diff", "spend_pct_diff",
        "impressions_QA", "impressions_Pipeline", "impressions_diff", "impressions_pct_diff",
    ]].sort_values(["status", "mmm_feature_og"], ignore_index=True).rename(
        columns=lambda c: c.replace("_QA", "_JS File").replace("_Pipeline", "_R File").replace("_pipeline", "_R File")
    )

    both = features["status"] == "Both"
    summary = {
        "Distinct features":                 features.shape[0],
        # "In both sources" is kept for the review-page card only; it is
        # skipped when writing the Summary sheet (see build_workbook).
        "In both sources":                   int(both.sum()),
        "feature not renamed":               int((features["match_type"] == "Auto").sum()),
        "feature renamed":                   int((features["match_type"] == "Manual").sum()),
        "Only in JS File":                   int((features["status"] == "Only in JS File").sum()),
        "Only in R File":                    int((features["status"] == "Only in R File").sum()),
        "Full match (spend & impressions)":  int(((features.loc[both, "spend_match"] == "Y") & (features.loc[both, "imp_match"] == "Y")).sum()),
        "Only Spend mismatch":               int((features.loc[both, "spend_match"] == "N").sum()),
        "Only Impressions mismatch":         int((features.loc[both, "imp_match"] == "N").sum()),
    }

    return features, summary

# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

# Sheet descriptions shown as a banner at the top of each tab.
DESC_SUMMARY = ("Summary of the MMM feature QA run: how the JavaScript-generated file (JS File) "
                "compares against the R-generated file (R File). The list below shows the client, "
                "the feature counts, and how many features matched, were renamed, or had value mismatches.")
DESC_JS      = ("Raw data loaded from the JavaScript-generated file (JS File), shown exactly as uploaded. "
                "This is one of the two inputs being compared.")
DESC_R       = ("Raw data loaded from the R-generated file (R File), shown exactly as uploaded. "
                "This is one of the two inputs being compared.")
DESC_COMP    = ("Feature-by-feature comparison of row counts, spend and impressions between the JS File and the R File. "
                "Columns are grouped into FEATURE, MATCH, ROWS, SPEND and IMPRESSIONS.   "
                "Row colours:  green = full match;  salmon = a value mismatch (the mismatched figures are highlighted in orange);  "
                "amber = the feature exists in only one file;  yellow (FEATURE cells) = a renamed feature — the R-file name and the "
                "JS-file name differ but were matched to the same feature, so both names are shown together.   "
                "The match_type column tells you how a feature pair was formed:  'Auto' = matched automatically because the names "
                "are the same once separators/case are ignored;  'Manual' = you mapped the rename yourself on the review screen.")

def build_workbook(qa_df, pipe_df, features, summary, client_name="") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    base_font  = Font(name="Calibri")
    bold       = Font(name="Calibri", bold=True)
    white_bold = Font(name="Calibri", bold=True, color="FFFFFF")
    bad_font   = Font(name="Calibri", bold=True, color="9C0006")
    title_font = Font(name="Calibri", bold=True, size=15)
    desc_font  = Font(name="Calibri", italic=True, size=12, color="404040")

    head_fill = PatternFill("solid", fgColor="D9E1F2")
    desc_fill = PatternFill("solid", fgColor="F2F2F2")
    wrap_top  = Alignment(wrap_text=True, vertical="top")
    center    = Alignment(horizontal="center", vertical="center")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left      = Alignment(horizontal="left", vertical="center")

    thin        = Side(style="thin", color="808080")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row colours
    ok_fill       = PatternFill("solid", fgColor="E2EFDA")   # light green  - full match
    mismatch_base = PatternFill("solid", fgColor="FCE4D6")   # light salmon - row has a mismatch
    only_fill     = PatternFill("solid", fgColor="FFF2CC")   # light amber  - only in one file
    accent_fill   = PatternFill("solid", fgColor="F4B084")   # rich orange  - the mismatched metric
    manual_fill   = PatternFill("solid", fgColor="FFE699")   # yellow tint  - renamed feature

    def _banner(target_ws, text, ncols, height=34):
        ncols = max(ncols, 1)
        target_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = target_ws.cell(row=1, column=1, value=text)
        c.font = desc_font; c.fill = desc_fill; c.alignment = wrap_top
        target_ws.row_dimensions[1].height = height

    def _write_df(target_ws, df):
        # No description banner on these raw-data tabs.
        # headers on row 1
        for ci, col in enumerate(df.columns, 1):
            cell = target_ws.cell(row=1, column=ci, value=col)
            cell.font = bold; cell.fill = head_fill
        # data from row 2
        for ri, vals in enumerate(df.values.tolist(), 2):
            for ci, val in enumerate(vals, 1):
                if isinstance(val, float) and pd.isna(val):
                    val = None
                cell = target_ws.cell(row=ri, column=ci, value=val)
                cell.font = base_font
                if isinstance(val, pd.Timestamp):
                    cell.number_format = "m/d/yyyy"
        for ci, col in enumerate(df.columns, 1):
            target_ws.column_dimensions[get_column_letter(ci)].width = max(len(str(col)), 18)
        target_ws.freeze_panes = "A2"
        # filter dropdowns on the header row (row 1) through the last data row
        if len(df.columns):
            last_row = len(df) + 1
            target_ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{last_row}"

    # --- JS File / R File raw sheets ---
    ws_qa   = wb.create_sheet(title="JS File", index=0)
    ws_pipe = wb.create_sheet(title="R File",  index=1)
    _write_df(ws_qa,   qa_df)
    _write_df(ws_pipe, pipe_df)

    # ----------------------------------------------------------------------
    # Summary
    # ----------------------------------------------------------------------
    ws.merge_cells("A1:B1")
    t = ws["A1"]; t.value = "MMM Feature QA"; t.font = title_font; t.fill = head_fill
    t.alignment = left
    ws.merge_cells("A2:B2")
    d = ws["A2"]; d.value = DESC_SUMMARY; d.font = desc_font; d.fill = desc_fill; d.alignment = wrap_top
    ws.row_dimensions[2].height = 110
    ws.append([])  # row 3 spacer

    # Summary list - no headers, Client name first, "In both sources" omitted.
    # Everything left aligned.
    rows_out = [("Client name", client_name)]
    rows_out += [(k, v) for k, v in summary.items() if k != "In both sources"]
    for label, value in rows_out:
        ws.append([label, value])
        r = ws.max_row
        lc = ws.cell(row=r, column=1); lc.font = bold;      lc.alignment = left
        vc = ws.cell(row=r, column=2); vc.font = base_font; vc.alignment = left
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18

    # ----------------------------------------------------------------------
    # Comparison  (categorised columns + rich mismatch highlighting)
    # ----------------------------------------------------------------------
    ws2 = wb.create_sheet("Comparison")

    # (label, band colour, header-tint colour, [column names])
    groups = [
        ("FEATURE",     "4472C4", "B4C6E7",
            ["feature_name_R_to_JS", "mmm_feature_og_JS File", "mmm_feature_og_R File"]),
        ("MATCH",       "7030A0", "CCC0DA",
            ["match_type", "spend_match", "imp_match"]),
        ("ROWS",        "548235", "C6E0B4",
            ["rows_JS File", "rows_R File"]),
        ("SPEND",       "BF8F00", "FFE699",
            ["spend_JS File", "spend_R File", "spend_diff", "spend_pct_diff"]),
        ("IMPRESSIONS", "808080", "D9D9D9",
            ["impressions_JS File", "impressions_R File", "impressions_diff", "impressions_pct_diff"]),
    ]
    display_cols = [c for _, _, _, cols in groups for c in cols]
    ncols = len(display_cols)
    idx_by_name = {name: i + 1 for i, name in enumerate(display_cols)}
    spend_group_idx   = [idx_by_name[c] for c in groups[3][3]]
    imp_group_idx     = [idx_by_name[c] for c in groups[4][3]]
    feature_group_idx = [idx_by_name[c] for c in groups[0][3]]

    # row 1: description banner (larger font, merged across all columns)
    _banner(ws2, DESC_COMP, ncols, height=70)

    # row 2: category band (merged, coloured, outlined)
    col = 1
    hdr_tint = {}
    for label, band_color, hdr_color, cols in groups:
        start, end = col, col + len(cols) - 1
        ws2.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
        band_fill = PatternFill("solid", fgColor=band_color)
        for ci in range(start, end + 1):
            bc = ws2.cell(row=2, column=ci)
            bc.fill = band_fill
            bc.border = cell_border        # outline the whole merged band
            hdr_tint[ci] = hdr_color
        top = ws2.cell(row=2, column=start)
        top.value = label
        top.font = white_bold
        top.alignment = center
        col = end + 1

    # row 3: column headers (tinted, outlined, centered, never wrapped)
    for ci, name in enumerate(display_cols, 1):
        hc = ws2.cell(row=3, column=ci, value=name)
        hc.font = bold
        hc.fill = PatternFill("solid", fgColor=hdr_tint[ci])
        hc.alignment = center             # no wrap -> names stay on one line
        hc.border = cell_border
    ws2.row_dimensions[3].height = 24

    # data rows from row 4
    r = 4
    for _, row in features.iterrows():
        spend_bad = row["spend_match"] == "N"
        imp_bad   = row["imp_match"] == "N"
        only_one  = row["status"] in ("Only in JS File", "Only in R File")
        is_manual = row.get("match_type") == "Manual"

        if only_one:
            base = only_fill
        elif spend_bad or imp_bad:
            base = mismatch_base
        else:
            base = ok_fill

        for ci, name in enumerate(display_cols, 1):
            val = row[name]
            if isinstance(val, float) and pd.isna(val):
                val = None
            cell = ws2.cell(row=r, column=ci, value=val)
            cell.fill = base
            cell.font = base_font
            cell.alignment = center
            if "pct" in name:
                cell.number_format = "0.00%"
            elif name.startswith("spend"):
                cell.number_format = "#,##0.00"

        # renamed feature -> tint the FEATURE cells yellow
        if is_manual:
            for ci in feature_group_idx:
                ws2.cell(row=r, column=ci).fill = manual_fill

        # rich accent on whichever metric actually mismatched (only meaningful
        # for genuine "Both" rows; an only-in-one-file row stays amber).
        if not only_one and spend_bad:
            for ci in spend_group_idx:
                c = ws2.cell(row=r, column=ci)
                c.fill = accent_fill
                if display_cols[ci - 1] in ("spend_diff", "spend_pct_diff"):
                    c.font = bad_font
        if not only_one and imp_bad:
            for ci in imp_group_idx:
                c = ws2.cell(row=r, column=ci)
                c.fill = accent_fill
                if display_cols[ci - 1] in ("impressions_diff", "impressions_pct_diff"):
                    c.font = bad_font
        r += 1

    # column widths - wide enough that no header name wraps
    for ci, name in enumerate(display_cols, 1):
        if name == "feature_name_R_to_JS":
            w = 48
        elif name.startswith("mmm_feature_og"):
            w = 36
        else:
            w = max(len(name) + 4, 13)
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.freeze_panes = "A4"   # keep description + category + header rows visible
    # Apply AutoFilter to the real header row (row 3) down through the last data
    # row. The merged banner (row 1) and merged category band (row 2) sit ABOVE
    # this range, so the filter range itself has no horizontal merges and the
    # dropdowns work. (Filtering failed before because no filter was set and the
    # merges above confused Excel's "apply filter" on the selected region.)
    last_row = r - 1
    if last_row >= 3:
        ws2.auto_filter.ref = f"A3:{get_column_letter(ncols)}{last_row}"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ---------------------------------------------------------------------------
# Suggestions for the review page
# ---------------------------------------------------------------------------

def build_review_data(features):
    """R File holds the originals; JS File holds the renamed versions.
    Anchor the mapping on R-file features and suggest a JS match by equal
    spend AND impressions. Returns (anchor_items, target_items, matched_items)
    where target_items are JS-only (rename candidates, eligible for
    auto-suggestion) and matched_items are JS features already matched in both
    files (selectable so an R-only feature can be SUMMED into one of them)."""
    js_only = features[features["status"] == "Only in JS File"].copy()
    r_only  = features[features["status"] == "Only in R File"].copy()
    both    = features[features["status"] == "Both"].copy()

    def _sig(spend, imp):
        # signature used to match a pair: spend to the cent, impressions to the unit
        return (round(float(spend), 2), round(float(imp)))

    # JS (renamed) features become the dropdown options
    target_items = [{
        "name":        row["mmm_feature_og"],
        "spend":       float(row["spend_JS File"]),
        "impressions": float(row["impressions_JS File"]),
        "rows":        int(row["rows_JS File"]),
    } for _, row in js_only.iterrows()]

    # JS features already present in BOTH files. Offering these as targets lets
    # an R-only feature be folded (summed) into a JS feature that another R
    # feature already auto-matched — e.g. R splits google_demand_gen into
    # dv360 + google_ads while the JS file keeps a single combined feature.
    matched_items = [{
        "name":        row["mmm_feature_og"],
        "spend":       float(row["spend_JS File"]),
        "impressions": float(row["impressions_JS File"]),
        "rows":        int(row["rows_JS File"]),
    } for _, row in both.iterrows()]

    # index JS features by (spend, impressions) for quick suggestion lookup
    js_by_sig = {}
    for it in target_items:
        js_by_sig.setdefault(_sig(it["spend"], it["impressions"]), []).append(it["name"])

    # index JS features by their normalized key for alias lookup
    js_by_key = {}
    for it in target_items:
        js_by_key.setdefault(_normalise_feature_key(it["name"]), []).append(it["name"])

    used = set()
    anchor_items = []
    for i, (_, row) in enumerate(r_only.iterrows()):
        spend = float(row["spend_R File"])
        imp   = float(row["impressions_R File"])
        r_name = row["mmm_feature_og"]
        r_key  = _normalise_feature_key(r_name)
        suggestion, basis = "", ""

        # 1) alias-known rename: a JS feature whose alias points at this R feature
        for js_key, names in js_by_key.items():
            if ALIAS_NORM.get(js_key) == r_key:
                for cand in names:
                    if cand not in used:
                        suggestion, basis = cand, "alias"
                        used.add(cand)
                        break
            if suggestion:
                break

        # 2) value-signature match: identical spend AND impressions
        if not suggestion:
            for cand in js_by_sig.get(_sig(spend, imp), []):
                if cand not in used:
                    suggestion, basis = cand, "values"
                    used.add(cand)
                    break

        anchor_items.append({
            "idx":         i,
            "name":        r_name,                  # R original
            "spend":       spend,
            "impressions": imp,
            "rows":        int(row["rows_R File"]),
            "suggestion":  suggestion,              # suggested JS rename
            "basis":       basis,                   # "alias" | "values" | ""
        })
    return anchor_items, target_items, matched_items

# ---------------------------------------------------------------------------
# Flask UI
# ---------------------------------------------------------------------------

PAGE = """
<!doctype html><title>MMM Feature QA</title>
<style>
body{font-family:system-ui;max-width:660px;margin:40px auto;padding:0 16px}
h1{margin-bottom:4px}
p.sub{color:#555;margin:0 0 20px}
input,button{font-size:15px;padding:8px;margin:4px 0}
input[type=text]{width:100%;box-sizing:border-box}
label{display:block;margin:10px 0 2px;font-weight:600}
.hint{font-size:13px;color:#666;margin:0 0 2px;font-weight:normal}
.section{background:#f7f8fc;border:1px solid #dde;border-radius:8px;padding:16px 20px;margin:12px 0}
button{background:#3b5bdb;color:#fff;border:none;border-radius:6px;padding:10px 24px;cursor:pointer;margin-top:12px}
button:hover{background:#2f4ac4}
.flash{background:#fee;border:1px solid #c00;border-radius:6px;padding:10px;margin:10px 0}
</style>
<h1>MMM Feature QA</h1>
<p class="sub">Compare the JAVA script generated file against a R script generated output. You'll get a chance to map renamed features before downloading a colour-coded Excel report.</p>
{% with msgs = get_flashed_messages() %}{% for m in msgs %}<div class="flash">{{ m }}</div>{% endfor %}{% endwith %}
<form method="post" enctype="multipart/form-data" action="/qa">
  <div class="section">
    <label>Client Name <span class="hint">(shown as the first row of the Summary tab)</span></label>
    <input type="text" name="client_name" placeholder="e.g. Acme Corp">
  </div>
  <div class="section">
    <label>JavaScript Generated File <span class="hint">(.xlsx or .csv)</span></label>
    <input type="file" name="qa_file" accept=".csv,.xlsx,.xls" required>
  </div>
  <div class="section">
    <label>R Script Generated File <span class="hint">(.xlsx or .csv)</span></label>
    <input type="file" name="pipe_file" accept=".csv,.xlsx,.xls" required>
  </div>
  <button type="submit">Review &amp; Map →</button>
</form>
"""

REVIEW = """
<!doctype html><title>Review &amp; Map renames</title>
<style>
body{font-family:system-ui;max-width:980px;margin:36px auto;padding:0 16px;color:#222}
h1{margin-bottom:2px}
p.sub{color:#555;margin:0 0 18px}
.cards{display:flex;gap:14px;flex-wrap:wrap;margin:0 0 22px}
.card{flex:1;min-width:150px;background:#f7f8fc;border:1px solid #dde;border-radius:8px;padding:12px 16px}
.card .n{font-size:24px;font-weight:700}
.card .l{font-size:13px;color:#555}
table{border-collapse:collapse;width:100%;margin:8px 0 24px;font-size:14px}
th,td{border:1px solid #e2e2ec;padding:7px 9px;text-align:left;vertical-align:top}
th{background:#d9e1f2}
td.num{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}
select{font-size:13px;padding:5px;max-width:340px}
.sugg{background:#fff7d6}
.tag{font-size:11px;font-weight:700;padding:1px 6px;border-radius:4px;margin-left:6px;color:#fff;text-transform:capitalize}
.tag.values{background:#2f9e44}.tag.alias{background:#7048e8}
.muted{color:#888}
.note{background:#eef4ff;border:1px solid #cdddff;border-radius:8px;padding:10px 14px;font-size:14px;margin:0 0 18px}
button{background:#3b5bdb;color:#fff;border:none;border-radius:6px;padding:11px 26px;cursor:pointer;font-size:15px}
button:hover{background:#2f4ac4}
a.back{display:inline-block;margin-left:14px;color:#3b5bdb;font-size:14px}
code{background:#f0f0f6;padding:1px 5px;border-radius:4px}
h2{font-size:17px;margin:26px 0 6px}
</style>
<h1>Review &amp; map renamed features</h1>
<p class="sub">Features below exist in only one file. If an R-file feature was <b>renamed</b> in the JS file, map it here so they collapse into a single comparison row.</p>

<div class="cards">
  <div class="card"><div class="n">{{ summary['In both sources'] }}</div><div class="l">In both</div></div>
  <div class="card"><div class="n">{{ anchor_items|length }}</div><div class="l">Only in R File</div></div>
  <div class="card"><div class="n">{{ target_items|length }}</div><div class="l">Only in JS File</div></div>
  <div class="card"><div class="n">{{ n_suggested }}</div><div class="l">Auto-suggested</div></div>
</div>

<form method="post" action="/download">
  <input type="hidden" name="token" value="{{ token }}">

  <div class="note">
    Suggested matches are <span class="sugg">pre-selected</span> — either because
    <b>spend and impressions both match</b> (<span class="tag values">values</span>)
    or a known <b>rename alias</b> (<span class="tag alias">alias</span>).
    The dropdown also lists features <b>already in both files</b>: pick one of those
    to <b>sum</b> this R feature into it (useful when the R side splits a feature
    that the JS side keeps combined). Leave a row on
    <code>(no match — keep separate)</code> to keep it as its own row.
  </div>

  {% if anchor_items %}
  <h2>Map "Only in R File" → "Only in JS File"</h2>
  <table>
    <tr><th style="width:34%">R File feature (original)</th><th class="num">R spend</th><th class="num">R impr.</th><th class="num">rows</th><th>maps to → JS File feature (renamed)</th></tr>
    {% for it in anchor_items %}
    <tr class="{{ 'sugg' if it.suggestion else '' }}">
      <td><code>{{ it.name }}</code>{% if it.basis %}<span class="tag {{ it.basis }}">{{ it.basis }}</span>{% endif %}</td>
      <td class="num">{{ '{:,.2f}'.format(it.spend) }}</td>
      <td class="num">{{ '{:,.0f}'.format(it.impressions) }}</td>
      <td class="num">{{ it.rows }}</td>
      <td>
        <input type="hidden" name="r_{{ it.idx }}" value="{{ it.name }}">
        <select name="map_{{ it.idx }}">
          <option value="">(no match — keep separate)</option>
          {% if target_items %}
          <optgroup label="Only in JS File (rename)">
          {% for t in target_items %}
          <option value="{{ t.name }}" {{ 'selected' if t.name == it.suggestion else '' }}>
            {{ t.name }} — {{ '{:,.2f}'.format(t.spend) }} / {{ '{:,.0f}'.format(t.impressions) }} impr.
          </option>
          {% endfor %}
          </optgroup>
          {% endif %}
          {% if matched_items %}
          <optgroup label="Already in both files (sum R feature into this)">
          {% for t in matched_items %}
          <option value="{{ t.name }}" {{ 'selected' if t.name == it.suggestion else '' }}>
            {{ t.name }} — {{ '{:,.2f}'.format(t.spend) }} / {{ '{:,.0f}'.format(t.impressions) }} impr.
          </option>
          {% endfor %}
          </optgroup>
          {% endif %}
        </select>
      </td>
    </tr>
    {% endfor %}
  </table>
  {% else %}
  <p class="muted">No R-only features to map.</p>
  {% endif %}

  {% if target_items %}
  <h2 class="muted">Reference: features only in JS File</h2>
  <table>
    <tr><th style="width:54%">JS File feature</th><th class="num">JS spend</th><th class="num">JS impr.</th><th class="num">rows</th></tr>
    {% for t in target_items %}
    <tr><td><code>{{ t.name }}</code></td><td class="num">{{ '{:,.2f}'.format(t.spend) }}</td><td class="num">{{ '{:,.0f}'.format(t.impressions) }}</td><td class="num">{{ t.rows }}</td></tr>
    {% endfor %}
  </table>
  {% endif %}

  <button type="submit">Generate QA report ↓</button>
  <a class="back" href="/">↺ Start over</a>
</form>
"""

@app.route("/")
def index():
    return render_template_string(PAGE)

@app.route("/qa", methods=["POST"])
def qa():
    qa_file   = request.files.get("qa_file")
    pipe_file = request.files.get("pipe_file")
    if not qa_file or not qa_file.filename:
        flash("Please upload the QA reference file."); return redirect(url_for("index"))
    if not pipe_file or not pipe_file.filename:
        flash("Please upload the pipeline output file."); return redirect(url_for("index"))
    try:
        qa_bytes   = qa_file.read()
        pipe_bytes = pipe_file.read()
        qa_sheet   = request.form.get("qa_sheet", "")
        pipe_sheet = request.form.get("pipe_sheet", "")
        client_name = request.form.get("client_name", "").strip()
        qa_df, pipe_df = prepare_dfs(qa_bytes, qa_file.filename, qa_sheet,
                                     pipe_bytes, pipe_file.filename, pipe_sheet)
        features, summary = compute_comparison(qa_df, pipe_df)
        anchor_items, target_items, matched_items = build_review_data(features)
    except Exception as e:
        flash(f"Error: {e}"); return redirect(url_for("index"))

    token = _cache_put({
        "qa_bytes": qa_bytes, "qa_name": qa_file.filename, "qa_sheet": qa_sheet,
        "pipe_bytes": pipe_bytes, "pipe_name": pipe_file.filename, "pipe_sheet": pipe_sheet,
        "client_name": client_name,
    })
    n_suggested = sum(1 for it in anchor_items if it["suggestion"])
    return render_template_string(REVIEW, token=token, summary=summary,
                                  anchor_items=anchor_items, target_items=target_items,
                                  matched_items=matched_items, n_suggested=n_suggested)

@app.route("/download", methods=["POST"])
def download():
    token = request.form.get("token", "")
    payload = SESSION_CACHE.get(token)
    if not payload:
        flash("Session expired — please upload the files again.")
        return redirect(url_for("index"))

    # Rebuild the manual map from the submitted dropdowns.
    # The anchor is the R-file (original) feature; the dropdown picks the JS
    # (renamed) feature. compute_comparison expects {JS name -> [R name, ...]}.
    # Several R rows may point at the SAME JS feature; collect them all so their
    # values get summed rather than the last one clobbering the rest.
    manual_map = {}
    for key, val in request.form.items():
        if key.startswith("map_") and val:          # val = JS (renamed) name
            idx = key[len("map_"):]
            r_name = request.form.get(f"r_{idx}", "")  # R (original) name
            if r_name:
                manual_map.setdefault(val, []).append(r_name)

    try:
        qa_df, pipe_df = prepare_dfs(
            payload["qa_bytes"], payload["qa_name"], payload["qa_sheet"],
            payload["pipe_bytes"], payload["pipe_name"], payload["pipe_sheet"],
        )
        features, summary = compute_comparison(qa_df, pipe_df, manual_map=manual_map)
        out = build_workbook(qa_df, pipe_df, features, summary,
                             client_name=payload.get("client_name", ""))
    except Exception as e:
        flash(f"Error: {e}"); return redirect(url_for("index"))

    client_name = payload.get("client_name", "").strip()
    safe_client = re.sub(r"[^\w\-]", "_", client_name).lower() if client_name else ""
    filename = f"{safe_client}_mmm_feature_qa_summary.xlsx" if safe_client else "mmm_feature_qa_summary.xlsx"
    return send_file(out, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    app.run(debug=True, port=5000)